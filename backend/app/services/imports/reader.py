"""Turn an uploaded workbook or CSV into plain rows, and typed values out of strings.

Everything here is deliberately forgiving about *shape* — stray whitespace, a BOM
from Excel, Arabic-Indic digits, a date typed three different ways — and completely
unforgiving about *meaning*. A file that is merely untidy should import; a file that
says something ambiguous should be refused with the row number.
"""

import csv
import io
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from app.core.exceptions import AppException

# Excel on an Arabic Windows writes these instead of 0-9. A spreadsheet that looks
# perfectly correct to the user would otherwise fail every numeric column.
_ARABIC_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩٫", "0123456789.")

# Dates people actually type, in the order they are tried. ISO first because the
# template asks for it and because it is the only one that is never ambiguous.
_DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y")

_TRUE = {"نعم", "yes", "y", "true", "1", "صح", "مفعل", "مُفعّل"}
_FALSE = {"لا", "no", "n", "false", "0", "خطأ", "معطل", "غير مفعل"}

MAX_ROWS_PER_SHEET = 50_000


class Cell:
    """A raw value plus where it came from, so every error can name a row."""

    __slots__ = ("sheet", "row", "column", "raw")

    def __init__(self, sheet: str, row: int, column: str, raw: str) -> None:
        self.sheet = sheet
        self.row = row
        self.column = column
        self.raw = raw


def normalise(value: object) -> str:
    """A cell as trimmed text, with the surprises taken out.

    Excel hands back datetimes, floats and None depending on the cell; the caller
    only ever wants a string it can parse deliberately. Floats are formatted without
    scientific notation, because `1e+15` in a barcode column is a lost barcode.
    """
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        # `%g` would give 1e+15; this keeps whole numbers whole and drops the
        # trailing zeros Excel invents.
        text = f"{value:.6f}".rstrip("0").rstrip(".")
    elif isinstance(value, Decimal):
        text = format(value, "f")
    else:
        text = str(value)
    # NFKC folds the Arabic presentation forms Excel sometimes emits; the BOM and
    # non-breaking spaces come from CSVs exported by Windows tools.
    text = unicodedata.normalize("NFKC", text)
    return text.replace("﻿", "").replace(" ", " ").strip()


def read_csv(content: bytes, sheet_name: str) -> list[dict[str, str]]:
    """One CSV file as a list of {column: text} rows."""
    text = _decode(content)
    reader = csv.DictReader(io.StringIO(text))
    if reader.fieldnames is None:
        raise AppException(400, f"الملف «{sheet_name}» فارغ أو لا يحتوي على صف عناوين.")
    headers = [normalise(h) for h in reader.fieldnames]
    rows: list[dict[str, str]] = []
    for raw in reader:
        # `None` key collects extra columns beyond the header; ignored on purpose so
        # a trailing comma does not sink the file.
        row = {
            header: normalise(raw.get(original, ""))
            for header, original in zip(headers, reader.fieldnames)
        }
        if any(row.values()):  # skip wholly blank lines
            rows.append(row)
        if len(rows) > MAX_ROWS_PER_SHEET:
            raise AppException(
                400,
                f"الملف «{sheet_name}» يتجاوز {MAX_ROWS_PER_SHEET:,} سطراً. "
                "قسّمه إلى ملفات أصغر.",
            )
    return rows


def read_workbook(content: bytes) -> dict[str, list[dict[str, str]]]:
    """An .xlsx as {sheet name: rows}. Sheets the spec does not know are ignored."""
    try:
        # `data_only` so formula cells yield their last computed value rather than
        # "=SUM(...)" — a file built by an accountant is full of them.
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as error:  # openpyxl raises a zoo of exception types
        raise AppException(
            400, "تعذّرت قراءة ملف Excel. تأكد أنه بصيغة .xlsx وغير تالف."
        ) from error

    sheets: dict[str, list[dict[str, str]]] = {}
    for worksheet in workbook.worksheets:
        iterator = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(iterator)
        except StopIteration:
            continue
        headers = [normalise(h) for h in header_row]
        if not any(headers):
            continue
        rows: list[dict[str, str]] = []
        for values in iterator:
            row = {
                header: normalise(value)
                for header, value in zip(headers, values)
                if header
            }
            if any(row.values()):
                rows.append(row)
            if len(rows) > MAX_ROWS_PER_SHEET:
                raise AppException(
                    400,
                    f"الورقة «{worksheet.title}» تتجاوز {MAX_ROWS_PER_SHEET:,} سطراً.",
                )
        sheets[normalise(worksheet.title)] = rows
    workbook.close()
    return sheets


def _decode(content: bytes) -> str:
    """Best-effort text out of an uploaded CSV.

    UTF-8 first, then the Windows Arabic code page, because "Save as CSV" in an
    Arabic Excel produces cp1256 and every product name would otherwise arrive as
    mojibake — silently, since mojibake is still valid text.
    """
    for encoding in ("utf-8-sig", "utf-8", "cp1256", "iso-8859-6"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise AppException(
        400, "تعذّر تحديد ترميز الملف. احفظه بصيغة CSV UTF-8 وأعد المحاولة."
    )


# --- Typed parsing. Each returns (value, error_message). ---
def parse_decimal(text: str) -> tuple[Decimal | None, str | None]:
    cleaned = text.translate(_ARABIC_DIGITS).replace(",", "").replace(" ", "")
    # A parenthesised number is how spreadsheets show negatives: (150.00).
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = "-" + cleaned[1:-1]
    if not cleaned:
        return None, None
    try:
        return Decimal(cleaned), None
    except InvalidOperation:
        return None, f"«{text}» ليس رقماً صحيحاً."


def parse_int(text: str) -> tuple[int | None, str | None]:
    value, error = parse_decimal(text)
    if error or value is None:
        return None, error
    if value != value.to_integral_value():
        return None, f"«{text}» يجب أن يكون رقماً صحيحاً بلا كسور."
    return int(value), None


def parse_date(text: str) -> tuple[date | None, str | None]:
    cleaned = text.translate(_ARABIC_DIGITS).strip()
    if not cleaned:
        return None, None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(cleaned, fmt).date(), None
        except ValueError:
            continue
    return None, f"«{text}» ليس تاريخاً مفهوماً. استخدم الصيغة YYYY-MM-DD."


def parse_bool(text: str) -> tuple[bool | None, str | None]:
    cleaned = text.strip().lower()
    if not cleaned:
        return None, None
    if cleaned in _TRUE:
        return True, None
    if cleaned in _FALSE:
        return False, None
    return None, f"«{text}» غير مفهوم. اكتب نعم أو لا."

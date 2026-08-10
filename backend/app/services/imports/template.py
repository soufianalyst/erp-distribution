"""The downloadable sample, generated from `spec.py` rather than maintained by hand.

A template that has drifted from the validator is worse than none: it is a promise
the importer then breaks, after the user has filled in ten thousand rows on the
strength of it. Generating both from one table makes drift impossible rather than
unlikely — and `test_imports.py` asserts the generated headers are exactly what the
parser requires, so even this file cannot quietly diverge.
"""

import csv
import io
import zipfile

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.imports.spec import GUIDE_RULES, SHEETS, Sheet

_HEADER_FILL = PatternFill("solid", fgColor="065F46")  # emerald-800
# Required columns get a distinct header so they are countable at a glance.
_REQUIRED_HEADER_FILL = PatternFill("solid", fgColor="92400E")  # amber-800
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=14, color="065F46")


def _autosize(worksheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def build_workbook() -> bytes:
    """The whole template as one .xlsx: a guide sheet plus one sheet per table."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    _build_guide(workbook)
    _build_examples(workbook)
    for sheet in SHEETS:
        _build_sheet(workbook, sheet)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_guide(workbook: Workbook) -> None:
    """Read first, and written so it can be read by someone who is not a programmer."""
    worksheet = workbook.create_sheet("دليل الاستخدام")
    worksheet.sheet_view.rightToLeft = True

    worksheet["A1"] = "استيراد البيانات من النظام القديم"
    worksheet["A1"].font = _TITLE_FONT
    row = 3

    worksheet.cell(row=row, column=1, value="قواعد عامة").font = Font(bold=True, size=12)
    row += 1
    for rule in GUIDE_RULES:
        worksheet.cell(row=row, column=1, value=f"• {rule}").alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        row += 1

    row += 1
    worksheet.cell(row=row, column=1, value="ترتيب التعبئة").font = Font(bold=True, size=12)
    row += 1
    for index, sheet in enumerate(SHEETS, start=1):
        worksheet.cell(row=row, column=1, value=f"{index}. {sheet.title} ({sheet.name})")
        worksheet.cell(row=row, column=2, value=sheet.purpose).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        row += 1

    row += 2
    worksheet.cell(row=row, column=1, value="شرح الأعمدة").font = Font(bold=True, size=12)
    row += 2
    for sheet in SHEETS:
        worksheet.cell(row=row, column=1, value=f"{sheet.title} — {sheet.name}").font = (
            Font(bold=True, color="065F46")
        )
        row += 1
        for header, text in (
            ("العمود", "A"), ("الوصف", "B"), ("إلزامي", "C"), ("النوع", "D"), ("ملاحظات", "E"),
        ):
            cell = worksheet.cell(row=row, column="ABCDE".index(text) + 1, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
        row += 1
        for column in sheet.columns:
            worksheet.cell(row=row, column=1, value=column.name)
            worksheet.cell(row=row, column=2, value=column.label)
            worksheet.cell(row=row, column=3, value="نعم" if column.required else "لا")
            kind = column.kind
            if column.kind == "choice":
                kind = " أو ".join(column.choices)
            worksheet.cell(row=row, column=4, value=kind)
            note = worksheet.cell(row=row, column=5, value=column.note)
            note.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        row += 1

    _autosize(worksheet, [26, 30, 10, 26, 70])


def _build_sheet(workbook: Workbook, sheet: Sheet) -> None:
    """A data sheet: one header row, and nothing else.

    The examples deliberately live on their own sheet rather than in this one. A
    template that ships filled-in demonstration rows relies on every user
    remembering to delete them, and the one who forgets imports a product called
    "أرز بسمتي 5 كجم" into their live catalogue. Nothing here can be uploaded by
    accident because there is nothing here to upload.

    The Arabic label and guidance ride along as cell comments, so the header row
    stays exactly what the parser expects while still being readable to someone who
    does not read the English column names.
    """
    worksheet = workbook.create_sheet(sheet.name)
    worksheet.sheet_view.rightToLeft = True

    for index, column in enumerate(sheet.columns, start=1):
        cell = worksheet.cell(row=1, column=index, value=column.name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL if not column.required else _REQUIRED_HEADER_FILL

        parts = [column.label, "" if not column.required else "(إلزامي)"]
        if column.choices:
            parts.append("القيم المسموحة: " + "، ".join(column.choices))
        if column.note:
            parts.append(column.note)
        cell.comment = Comment("\n".join(p for p in parts if p), "ERP")

    worksheet.freeze_panes = "A2"
    _autosize(worksheet, [max(14, min(30, len(c.name) + 6)) for c in sheet.columns])


def _build_examples(workbook: Workbook) -> None:
    """Every table filled in, on one sheet, to copy the shape from — never uploaded."""
    worksheet = workbook.create_sheet("أمثلة")
    worksheet.sheet_view.rightToLeft = True
    worksheet["A1"] = "أمثلة توضيحية — هذه الورقة للاطلاع فقط ولا تُقرأ عند الاستيراد"
    worksheet["A1"].font = _TITLE_FONT

    row = 3
    for sheet in SHEETS:
        worksheet.cell(row=row, column=1, value=f"{sheet.title} ({sheet.name})").font = (
            Font(bold=True, size=12, color="065F46")
        )
        row += 1
        for index, column in enumerate(sheet.columns, start=1):
            cell = worksheet.cell(row=row, column=index, value=column.name)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL if not column.required else _REQUIRED_HEADER_FILL
        row += 1
        for offset in range(2):
            for index, column in enumerate(sheet.columns, start=1):
                worksheet.cell(row=row, column=index, value=column.examples[offset])
            row += 1
        row += 2

    _autosize(worksheet, [max(14, min(30, 20)) for _ in range(14)])


def build_csv(sheet: Sheet, *, with_examples: bool = False) -> bytes:
    """One sheet as a standalone CSV, for anyone not using Excel.

    UTF-8 with a BOM: without it Excel opens the file in the local code page and
    every Arabic name arrives as mojibake, which users reasonably read as our bug.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([column.name for column in sheet.columns])
    if with_examples:
        for offset in range(2):
            writer.writerow([column.examples[offset] for column in sheet.columns])
    return buffer.getvalue().encode("utf-8-sig")


def build_csv_bundle() -> bytes:
    """Every sheet as a CSV, plus the rules as a text file, in one zip."""
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        for sheet in SHEETS:
            # The upload-ready file is empty below the header; the filled one sits in
            # a folder the importer will never be pointed at.
            archive.writestr(f"{sheet.name}.csv", build_csv(sheet))
            archive.writestr(
                f"أمثلة/{sheet.name}.csv", build_csv(sheet, with_examples=True)
            )
        guide = ["استيراد البيانات من النظام القديم", "", "قواعد عامة:"]
        guide += [f"  - {rule}" for rule in GUIDE_RULES]
        guide += ["", "ترتيب التعبئة:"]
        for index, sheet in enumerate(SHEETS, start=1):
            guide.append(f"  {index}. {sheet.name} — {sheet.title}")
            guide.append(f"     {sheet.purpose}")
            for column in sheet.columns:
                flag = "إلزامي" if column.required else "اختياري"
                line = f"       • {column.name} ({column.label}) — {flag}"
                if column.note:
                    line += f" — {column.note}"
                guide.append(line)
            guide.append("")
        archive.writestr("اقرأني.txt", "\n".join(guide).encode("utf-8"))
    return buffer.getvalue()
